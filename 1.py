from openpyxl import Workbook
import pandas as pd

df = pd.read_csv('score.txt',header = None)
fn = 'score.xlsx'
wb = Workbook()
ws = wb.create_sheet(title='score')

py1 =sorted(list((df[1])))
py1.pop(0)
py1.pop(-1)

py2 =sorted(list((df[2])))
py2.pop(0)
py2.pop(-1)

py3 =sorted(list((df[3])))
py3.pop(0)
py3.pop(-1)

sc ={1:sum(py1)/8,2:sum(py2)/8,3:sum(py3)/8}
rs = max(sc.items(),key =lambda x:x[1])

ws['A1'] = rs[0]
ws['B1'] = rs[1]
 
wb.save(fn)



